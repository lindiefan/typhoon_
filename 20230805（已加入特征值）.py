"""
项目开发日记
前一段停用词表和情感词典的更新

0714
就是一共是三个工作：
1. 进行高频词汇筛查，准备做台风主题词典
2. 对时间进行处理，处理为天+小时的形式（分为两列）
3. 对地点进行处理，整理为全部都是城市具体数据的、有省份数据的、还有原来版本的（采用多版本来训练）

0716
我把计算规则那边改了一点，就改了一个是两个情绪词之间范围太大的问题，我把所有规则都改成紧邻（比如说双重否定要求两个否定词都和情绪词紧靠）
还有改了一下权重，之前乘0.8改成0.5,1.2改成2

0723
加入了特征值们，可以对特征值进行赋值了，形成了特征值表格（标注由我标注的，加上GPT协助）

0724
大概改了3个地方：
1、副词乘以全句改为副词乘以紧连其后的情绪词（若不紧连则视为无效）
2、改掉了”副词+否定词+情绪词“的连乘bug（也是仅限于3个词紧连有效）
3、修改了计算句末感叹号倍数的部分（原本的乘法次数有误）
4. 在jiaba包加载了新的词典————流行词词典和天气专用词汇词典，都经过去重处理，
来源于搜狗细胞词典的网络流行新词词典和诸多天气、气象词汇词典，将scel转换为txt，经过复制后去重得到台风专题词表。

0805
处理好了emoji
1.准备先识别微博的官方的emoji，然后将emoji按积极、中性消极分（或者可以求和，20来个没有的按0.25 0.25 0.5 最后得到0.25来分配也可以） 微博emoji来源于微博表情（即emojiall的微博专题）
2.而后识别文本中的emoji，计分或者计算情感值后在计算也可以。
3. 处理emoji时，很多emoji是两个字节单位，也有不少不是的，需要慎重处理，对emoji的处理中，对大部分按两个字节处理，对剩下的用手工标注后，蛮多emoji是多个基础emoji结合而成，这种就要数有多少个
基础emoji，然后数量*2去处理。

计分 求最高 求平均 取分类

"""
import jieba.analyse
import openpyxl
import numpy as np
import xlwt
import pandas as pd
# from ltp import StnSplit
from pyltp import SentenceSplitter  #
from openpyxl import load_workbook

def run_score(contents):
    scores_list = []
    feature_list=[]
    dili_eng = []
    typhoon_eng = []
    feeling_eng = []
    do_eng = []
    influence_eng = []
    damage_eng = []
    wind_eng = []
    rain_eng = []
    cn = 0
    for content in contents:
        if content != '':

            # if emoji_list[cn] != ' ' and emoji_list[cn] in emoji_word:
            #     index1 = emoji_word.index(emoji_list[cn])
            #     score = single_sentiment_score(content) + 3 * emoji_emotional[index1] # 三是权重
            score,dili, typhoon, feeling, do, influence, damage, wind, rain = single_sentiment_score(content)  # 对每条微博调用函数求得打分
            scores_list.append((score, content))  # 形成（分数，微博）元组
            dili_eng.append(dili)
            typhoon_eng.append(typhoon)
            feeling_eng.append(feeling)
            do_eng.append(do)
            influence_eng.append(influence)
            damage_eng.append(damage)
            wind_eng.append(wind)
            rain_eng.append(rain)
    return scores_list,dili_eng, typhoon_eng, feeling_eng,do_eng, influence_eng, damage_eng,wind_eng, rain_eng


def single_sentiment_score(text_sent):

    dili_eng = 0
    typhoon_eng =0
    feeling_eng =0
    do_eng =0
    influence_eng = 0
    damage_eng =0
    wind_eng =0
    rain_eng =0
    sentiment_scores = []
    emoji_score = []

    # 对单条微博分句
    sentences = cut_sentence(text_sent)
    # print(sentences)
    for sent in sentences:

        # 查看分句结果
        # print('分句：',sent)
        # 分词
        words = jieba.lcut(sent)
        seg_words = del_stopwords(words)
        # 查看分词结果
        # print(seg_words)
        # i，s 记录情感词和程度词出现的位置
        i = 0  # 记录扫描到的词位子
        s = 0  # 记录情感词的位置
        sentiment_score=0  # 单句情绪值得分初始化
        # l_adv = []  # 修饰全句的副词权重记录表初始化
        # 逐个查找情感词
        for word in seg_words:
            # emoji查找
            if word in weibo_emojiname:
                weibo_emoji_score_temp = emoji_weibo_emotion_score[weibo_emojiname.index(word)]
                weibo_emoji_emotion_category_temp = emoji_weibo_emotion_category[weibo_emojiname.index(word)]
                """
                # 这里后续还要把计算方式写好，是要求和(对短文本)还是计分，还是乘上去（超级大 超级小）
                """
            if word in emoji_all:
                emoji_all_score_temp = emoji_all_emotion_score[emoji_all.index(word)]
                emoji_all_emotion_category_temp = emoji_all_emotion_category[emoji_all.index(word)]
                """
                这里同上 需要做有一定的处理
                """
            # 特征值赋值操作
            if word in dili_list:
                dili_eng_temp = dili_eng_list[dili_list.index(word)]
                if dili_eng_temp > dili_eng:
                    dili_eng = dili_eng_temp
            if word in typhoon_list:
                typhoon_eng_temp = typhoon_eng_list[typhoon_list.index(word)]
                if typhoon_eng_temp > typhoon_eng:
                    typhoon_eng = typhoon_eng_temp
            if word in feeling_list:
                feeling_eng_temp = feeling_eng_list[feeling_list.index(word)]
                if feeling_eng_temp > feeling_eng:
                    feeling_eng = feeling_eng_temp
            if word in do_list:
                do_eng_temp = do_eng_list[do_list.index(word)]
                if do_eng_temp > do_eng:
                    do_eng = do_eng_temp
            if word in influence_list:
                influence_eng_temp = influence_eng_list[influence_list.index(word)]
                if influence_eng_temp > influence_eng:
                    influence_eng = influence_eng_temp
            if word in damage_list:
                damage_eng_temp = damage_eng_list[damage_list.index(word)]
                if damage_eng_temp > damage_eng:
                    damage_eng = damage_eng_temp
            if word in strength_list:
                wind_eng_temp = strength_eng_list[strength_list.index(word)]
                if wind_eng_temp > wind_eng:
                    wind_eng = wind_eng_temp
            if word in rain_list:
                rain_eng_temp = rain_eng_list[rain_list.index(word)]
                if rain_eng_temp > rain_eng:
                    rain_eng = rain_eng_temp

            # 开始进行情感词处理
            # print(word,poscount,1)
            if word in posiword_list or word in nageword_list:
                if word in posiword_list:
                    s1=1
                else:
                    s1=-1
                # print(f'{word}的初始得分是{s1}')
                # 在情感词前面寻找程度副词和否定词
                sign=0  #否定词个数初始化
                for w in seg_words[s:i]:
                    if w in inversedict:
                        sign+=1  #记录否定词的个数
                if sign==0:  # 情绪词前无否定词时，直接将副词权重乘以该情绪词得分
                    s1=match_adverb(seg_words[i-1],s1)
                elif sign==1:
                    for x in range(s,i+1):
                        if seg_words[x] in inversedict:  # 找到否定词的位置x
                            break
                    if x+2==i:   # 否定词+程度副词+情绪词
                        if w in mostdict or w in verydict or w in moredict:
                            s1*=0.5  # 否定词+高或较高程度副词+情绪词=情感减弱
                        elif w in ishdict or w in insufficientdict:
                            s1 *= (-0.5)  # 否定词+低或较低程度副词+情绪词=情感取反再减弱
                    elif x+1==i:  # 程度副词+否定词+情绪词
                        s1 = match_adverb(seg_words[x-1],s1)  #副词+否定词+情绪词=副词权重*情绪词得分*(-1)
                        s1*=(-1)
                elif sign==2:
                    for x in range(s,i+1):
                        if seg_words[x] in inversedict:  # 找到第一个否定词的位置x
                            break
                    if x+2==i and seg_words[x+1] in inversedict:
                        s1*=2  #双重否定 情绪增强
                s = i + 1  # 记录情感词位置
                # print(f'{word}的加权得分是{s1}')
                sentiment_score+=s1
            i += 1  # 定位情感词的位置
        if sent.count('!')+sent.count('！')==1:
            sentiment_score *= 2  # 对整句话情绪值的得分放大
        elif sent.count('!')+sent.count('！')>1:
            if sentiment_score == 0:  # 如果一个句子当中没有情绪词的话，多重感叹号取“震惊”（负面）义，强度3
                sentiment_score -= 3
            else:
                sentiment_score *= 4  # 对整个分句的得分放大
        # 计算情感值
        # print(poscount,5)
        sentiment_scores.append(sentiment_score)
        # 查看每一句的情感值
        # print('分句分值：',sentiment_score)
    sentiment_sum = 0
    for s in sentiment_scores:
        # 计算出一条微博的总得分
        sentiment_sum += s
    return sentiment_sum,dili_eng, typhoon_eng, feeling_eng,do_eng, influence_eng, damage_eng,wind_eng, rain_eng

def cut_sentence(text):
    sentences = SentenceSplitter.split(text)  #
    # sentences = StnSplit().split(text)
    sentence_list = [ w for w in sentences]
    return sentence_list

#去停用词函数
def del_stopwords(words):
    # 读取停用词表
    # stopwords = read_file(r"C:\Users\lbq-1\Desktop\social big data\stopwords.txt") #这个用起来感觉还行 试试看 三个停用词表合起来的效果
    stopwords = read_file(r"C:\Users\lbq-1\Desktop\social big data\new停用词表.txt")
    # 去除停用词后的句子
    new_words = []
    for word in words:
        if word not in stopwords:
            new_words.append(word)
    return new_words

#读取文件，文件读取函数
def read_file(filename):
    with  open(filename,'r',encoding='utf-8')as f:
        text = f.read()
        #返回list类型数据
        text = text.split('\n')
    return text

#程度副词处理，对不同的程度副词给予不同的权重
# def match_adverb(word,sentiment_value):
#
#     #最高级权重为
#     if word in abverb_word:
#         index = abverb_word.index(word)
#         sentiment_value = sentiment_value * float(abverb_weight[index])
#         print(abverb_weight[index])
#         print("sentiment_value",sentiment_value)
#     # #比较级权重
#     # elif word in verydict:
#     #     sentiment_value *= 6
#     # #比较级权重
#     # elif word in moredict:
#     #     sentiment_value *= 4
#     # #轻微程度词权重
#     # elif word in ishdict:
#     #     sentiment_value *= 2
#     # #相对程度词权重
#     # elif word in insufficientdict:
#     #     sentiment_value *= 0.5
#     #否定词权重============================pay a attention to here!!!
#     # elif word in inversedict:
#     #     sentiment_value *= -1
#     else:
#         sentiment_value *= 1
#     return sentiment_value
# 获取六种权值的词，根据要求返回list，这个函数是为了配合Django的views下的函数使用
def weighted_value(request):
    result_dict = []
    if request == "one":
        result_dict = read_file(r"C:\Users\lbq-1\Desktop\social big data\chinese_sentiment_dictionary-master\chinese_sentiment_dictionary-master\file\emotion dictionary\abverb\most.txt")
    elif request == "two":
        result_dict = read_file(r"C:\Users\lbq-1\Desktop\social big data\chinese_sentiment_dictionary-master\chinese_sentiment_dictionary-master\file\emotion dictionary\abverb\very.txt")
    elif request == "three":
        result_dict = read_file(r"C:\Users\lbq-1\Desktop\social big data\chinese_sentiment_dictionary-master\chinese_sentiment_dictionary-master\file\emotion dictionary\abverb\more.txt")
    elif request == "four":
        result_dict = read_file(r"C:\Users\lbq-1\Desktop\social big data\chinese_sentiment_dictionary-master\chinese_sentiment_dictionary-master\file\emotion dictionary\abverb\ish.txt")
    elif request == "five":
        result_dict = read_file(r"C:\Users\lbq-1\Desktop\social big data\chinese_sentiment_dictionary-master\chinese_sentiment_dictionary-master\file\emotion dictionary\abverb\insufficiently.txt")
    elif request == "six":
        result_dict = read_file(r"C:\Users\lbq-1\Desktop\social big data\chinese_sentiment_dictionary-master\chinese_sentiment_dictionary-master\file\emotion dictionary\abverb\no.txt")
    # elif request == 'posdict':
    #     result_dict = read_file(r"E:\学习笔记\NLP学习\NLP code\情感分析3\emotion_dict\pos_all_dict.txt")
    # elif request == 'negdict':
    #     result_dict = read_file(r"E:\学习笔记\NLP学习\NLP code\情感分析3\emotion_dict\neg_all_dict.txt")
    else:
        pass
    return result_dict
#

#程度副词处理，对不同的程度副词给予不同的权重
def match_adverb(word,sentiment_value):
    #最高级权重为
    if word in mostdict:
        sentiment_value *= 8
    #比较级权重
    elif word in verydict:
        sentiment_value *= 6
    #比较级权重
    elif word in moredict:
        sentiment_value *= 4
    #轻微程度词权重
    elif word in ishdict:
        sentiment_value *= 2
    #相对程度词权重
    elif word in insufficientdict:
        sentiment_value *= 0.5
    #否定词权重
    #elif word in inversedict:
    #    sentiment_value *= -1
    else:
        sentiment_value *= 1
    return sentiment_value
if __name__ == '__main__':
    print('Processing........')
    data_feature = pd.read_excel(r"C:\Users\lbq-1\Desktop\social big data\research data\特征值表格.xlsx")
    dili_list = data_feature['地理灾害'].values.tolist()
    dili_eng_list = data_feature['feature_dili'].values.tolist()

    typhoon_list = data_feature['台风强度'].values.tolist()
    typhoon_eng_list = data_feature['feature_typhoon'].values.tolist()

    feeling_list = data_feature['台风给人的感受'].values.tolist()
    feeling_eng_list = data_feature['feature_human_feeling'].values.tolist()

    do_list = data_feature['台风行为'].values.tolist()
    do_eng_list = data_feature['feature_typhoon_do'].values.tolist()

    influence_list = data_feature['给人的影响'].values.tolist()
    influence_eng_list = data_feature['feature_human_influence'].values.tolist()

    damage_list = data_feature['经济损失'].values.tolist()
    damage_eng_list = data_feature['feature_economic_damage'].values.tolist()

    strength_list = data_feature['风力强度'].values.tolist()
    strength_eng_list = data_feature['feature_wind_strength'].values.tolist()

    rain_list = data_feature['雨量大小'].values.tolist()
    rain_eng_list = data_feature['feature_rain_strength'].values.tolist()
    # 读取数据并清洗

    contentdf = pd.read_excel(
        r"C:\Users\lbq-1\Desktop\social big data\research data\city_test2.xlsx")  # 原来是"C:\Users\lbq-1\Desktop\social big data\微博内容new.csv"
    # contentdf = pd.read_csv(r"C:\Users\lbq-1\Desktop\social big data\try.csv")
    # commentdf = pd.read_csv(r"C:\Users\lbq-1\Desktop\social big data\微博评论0228.csv")
    cidiandf_posi = read_file(r"C:\Users\lbq-1\Desktop\social big data\new_positive_dictionary.txt")
    cidiandf_nega = read_file(r"C:\Users\lbq-1\Desktop\social big data\new_negative_dictionary.txt")
    print("加载词库ing")
    jieba.load_userdict(r"C:\Users\lbq-1\Desktop\social big data\research data\网络流行新词2【官方推荐】.txt")
    jieba.load_userdict(r"C:\Users\lbq-1\Desktop\social big data\research data\气象词汇词库3.txt")
    jieba.load_userdict(r"C:\Users\lbq-1\Desktop\social big data\research data\weibo_emoji_dictionary.txt")
    jieba.load_userdict(r"C:\Users\lbq-1\Desktop\social big data\research data\all_emoji_dictionary_new.txt")
    print("加载词库over")
    emoji_weibo = pd.read_excel(r"C:\Users\lbq-1\Desktop\social big data\weibo_emoji url and emotion.xlsx")
    emoji_emotion = pd.read_excel(r"C:\Users\lbq-1\Desktop\social big data\emoji all emotion_aaa.xlsx")
    weibo_emojiname = emoji_weibo['emojiname']
    emoji_weibo_emotion_score = emoji_weibo['emotion_add'].values.tolist()
    emoji_weibo_emotion_category = emoji_weibo["emotion_category"].values.tolist()
    print(weibo_emojiname[0],type(weibo_emojiname[0]),emoji_weibo_emotion_category[0],emoji_weibo_emotion_score[0],type(emoji_weibo_emotion_score[0]))
    emoji_all = emoji_emotion['Emoji']
    emoji_all_text = emoji_emotion['Text']
    emoji_all_emotion_score = emoji_emotion['emotion_add'].values.tolist()
    emoji_all_emotion_category = emoji_emotion['emotion_category'].values.tolist()
    print(type(emoji_all[0]))
    # no = read_file(r"C:\Users\lbq-1\Desktop\social big data\chinese_sentiment_dictionary-master\chinese_sentiment_dictionary-master\file\emotion dictionary\abverb\no.txt")
    # 处理 emoji词典 暂时以1 0 -1 处理
    #emoji_dictory = pd.read_excel(r"C:\Users\lbq-1\Desktop\social big data\emoji url and emotion.xlsx")
    #emoji_word = emoji_dictory["emoji_word"].values.tolist()
    #emoji_negative = emoji_dictory["negative"].values.tolist()
    #emoji_normal =emoji_dictory["normal"].values.tolist()
    #emoji_positive =emoji_dictory["positive"].values.tolist()
    #emoji_emotional = []
   # for i in range(len(emoji_word)):
    #    if emoji_positive[i] > emoji_normal[i] and emoji_positive[i] > emoji_negative[i]:
    #        emoji_emotional.append(1)
    #    if emoji_negative[i] > emoji_normal[i] and emoji_negative[i] > emoji_positive[i]:
    #        emoji_emotional.append(-1)
    #    if emoji_normal[i] > emoji_positive[i] and emoji_normal[i] > emoji_negative[i]:
    #        emoji_emotional.append(0)
    print("reading sentiment dict .......")
    # 读取情感词典
    # posdict = weighted_value('posdict')
    # negdict = weighted_value('negdict')
    # 读取程度副词词典
    # 权值为2
    mostdict = weighted_value('one')
    # 权值为1.75
    verydict = weighted_value('two')
    # 权值为1.50
    moredict = weighted_value('three')
    # 权值为1.25
    ishdict = weighted_value('four')
    # 权值为0.25
    insufficientdict = weighted_value('five')
    # 权值为-1
    inversedict = weighted_value('six')
    # f = open('no4.txt','w')
    # print(no)
    # for i in no:
    #     print(i)
    #     e = i.replace('\t','')
    #     e = e + '\n'
    #     f.write(e)
    # f.close()
    # ajkd
    # ciidandf = pd.read_csv(r"C:\Users\lbq-1\Desktop\social big data\情感词汇.csv")
    # cidiandf_posi = read_file(r"C:\Users\lbq-1\Desktop\social big data\chinese_sentiment_dictionary-master\chinese_sentiment_dictionary-master\file\情感词典\综合版本\综合情感词典（积极）.txt")
    # cidiandf_nega = read_file(r"C:\Users\lbq-1\Desktop\social big data\chinese_sentiment_dictionary-master\chinese_sentiment_dictionary-master\file\情感词典\综合版本\综合情感词典（消极）.txt")
    # cidiandf_posi1 = set(cidiandf_posi)
    # print(cidiandf_posi)
    # cidiandf_nega1 = set(cidiandf_nega)
    # cidiandf_posi = list(cidiandf_posi1)
    # cidiandf_nega = list(cidiandf_nega1)
    # f_posi = open('negative_dictionary',mode='w')
    # for i in cidiandf_nega:
    #     if i == '':
    #         pass
    #     e = i.replace(' ','')+'\n'
    #     print(e)
    #     f_posi.write(e)
    # f_posi.close()
    #
    #
    # print(cidiandf_posi)
    content_list = contentdf['text'].values.tolist()
    #emoji_list = contentdf['emoji'].values.tolist()
    #print(emoji_list)
    #emojiimage_list = contentdf['emoji_image'].values.tolist()
    bowentime_list = contentdf['created_at'].values.tolist()
    contentdf['user_location'].fillna('缺失', inplace=True)
    bowenplace_list = contentdf['user_province'].values.tolist()
    bowenplace_list = contentdf['user_location'].values.tolist()
    # ee = []
    # for i in bowenplace_list:
    #     e = str(i[0:2])
    #     ee.append(e)

    # import xlwt
    #
    # workbook = xlwt.Workbook(encoding='utf-8')
    # worksheet = workbook.add_sheet('sheet1')
    # firstline = ['字母英文', '首尾相同情况', '重复字母数量', '首字母出现频率', '字母出现频率', '含义', '词性']
    # for i in range(33791):
    #     i = i + 1
    #     worksheet.write(i, 1, bowenplace_list[i - 1])
    #     worksheet.write(i, 2, ee[i - 1])
    # workbook.save("2千词6项.xls")
    # kgkjhkgjgj
    username_list = contentdf['user_screen_name'].values.tolist()
    posiword_list = cidiandf_posi
    nageword_list = cidiandf_nega
    # adverb = read_file(r"C:\Users\lbq-1\Desktop\social big data\程度副词.txt")

    #地点数据清洗
    # city = read_file(r"C:\Users\lbq-1\Desktop\social big data\中国地级市.txt")
    # city1 = []
    # for i in city:
    #     i = i.replace('市','').replace('自治区','').replace('地区','').replace('自治州','').replace('','')
    #     city1.append(i)
    # didianbaoliu = []
    # for i in bowenplace_list:
    #     index = 0
    #     if i != '缺失':
    #         for j in city1:
    #             if i.find(j) != -1:
    #                 index = 1
    #     didianbaoliu.append(index)
    # for i in range(1000,1500):
    #     print(bowenplace_list[i],didianbaoliu[i])
    # import xlwt
    #
    # workbook = xlwt.Workbook(encoding='utf-8')
    # worksheet = workbook.add_sheet('sheet1')
    # for i in range(33791):
    #     i = i + 1
    #     worksheet.write(i, 1, content_list[i - 1])
    #     worksheet.write(i, 2, didianbaoliu[i - 1])
    #     worksheet.write(i, 3, bowenplace_list[i - 1])
    # workbook.save("dianle.xls")
    # firstline = ['字母英文', '首尾相同情况', '重复字母数量', '首字母出现频率', '字母出现频率', '含义', '词性']
    print("=======================")
    # abverb_word = []
    # abverb_weight = []
    # for i in adverb:
    #     kk = i.split(',')
    #     abverb_word.append(kk[0])
    #     abverb_weight.append(kk[1])
    # print(len(abverb_word))
    # print(len(abverb_weight))

    # 构建台风情感词典——寻找高频词
    # print(content_list[0])
    # print(type(content_list[0]))
    # content_list_str = ' '.join(str(i) for i in content_list)
    # print(content_list_str)
    # keywords_top25 = jieba.analyse.extract_tags(content_list_str, withWeight=True, topK=4000)
    # import xlwt
    #
    # workbook = xlwt.Workbook(encoding='utf-8')
    # worksheet = workbook.add_sheet('sheet1')
    # for i in range(4000):
    #     i = i + 1
    #     worksheet.write(i, 1, keywords_top25[i - 1][0])
    #     worksheet.write(i, 2, keywords_top25[i - 1][1])
    #     # worksheet.write(i, 3, content_list[i - 1])
    # workbook.save("gaopincihui.xls")
    # print(keywords_top25)

    scores,dili_eng, typhoon_eng, feeling_eng,do_eng, influence_eng, damage_eng,wind_eng, rain_eng = run_score(content_list)

    print("喜大普奔！！！！！")
    # print(scores[0:100])
    # print(features[0:100])
    emotion_score=[]
    for score in scores:
        # print('情感分值：', score[0])
        es = score[0]
        emotion_score.append(es)
    contentdf['emotion_score'] = emotion_score
    contentdf["地理灾害"] = dili_eng
    contentdf["台风强度"] = typhoon_eng
    contentdf["台风给人的感受"] =feeling_eng
    contentdf["台风行为"] = do_eng
    contentdf["给人的影响"] = influence_eng
    contentdf["经济损失"] = damage_eng
    contentdf["风力强度"] = wind_eng
    contentdf["雨量大小"] = rain_eng
    contentdf.to_excel(r"C:\Users\lbq-1\Desktop\social big data\city_test_result.xlsx")

    # import xlwt
    #
    # workbook = xlwt.Workbook(encoding='utf-8')
    # worksheet = workbook.add_sheet('sheet1')
    # worksheet.write(0, 1, "content_list")
    # worksheet.write(0, 2, "emotion_score")
    # worksheet.write(0, 3, "地理灾害")
    # worksheet.write(0, 4, '台风强度')
    # worksheet.write(0, 5,  '台风给人的感受')
    # worksheet.write(0, 6,  '台风行为')
    # worksheet.write(0, 7,  '给人的影响')
    # worksheet.write(0, 8,  '经济损失')
    # worksheet.write(0, 9,  '风力强度')
    # worksheet.write(0, 10, '雨量大小')
    # for i in range(len(username_list)):
    #     i = i + 1
    #     worksheet.write(i, 1, content_list[i - 1])
    #     worksheet.write(i, 2, emotion_score[i - 1])
    #     worksheet.write(i, 3, features[i - 1][0])
    #     worksheet.write(i, 4, features[i - 1][1])
    #     worksheet.write(i, 5, features[i - 1][2])
    #     worksheet.write(i, 6, features[i - 1][3])
    #     worksheet.write(i, 7, features[i - 1][4])
    #     worksheet.write(i, 8, features[i - 1][5])
    #     worksheet.write(i, 9, features[i - 1][6])
    #     worksheet.write(i, 10, features[i - 1][7])
    #     # worksheet.write(i, 3, content_list[i - 1])
    # workbook.save("city计算结果.xls")
